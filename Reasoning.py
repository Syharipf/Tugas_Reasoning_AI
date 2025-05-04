from openpyxl import load_workbook, Workbook

# ================= FUNGSI KEANGGOTAAN FUZZY =================
# Fungsi untuk menentukan derajat keanggotaan variabel SERVIS (1-100)
def servis_membership(x):
    return {
        # Fungsi segitiga turun untuk 'buruk' (50 → 0)
        'buruk': max(min((50 - x) / 30, 1), 0),
        # Fungsi segitiga untuk 'sedang' (40 → 80)
        'sedang': max(min((x - 40) / 20, (80 - x) / 20), 0),
        # Fungsi segitiga naik untuk 'baik' (60 → 100)
        'baik': max(min((x - 60) / 30, 1), 0)
    }

# Fungsi untuk menentukan derajat keanggotaan variabel HARGA (Rp 25.000-55.000)
def harga_membership(x):
    return {
        # Fungsi segitiga turun untuk 'murah' (30k → 0)
        'murah': max(min((30000 - x) / 10000, 1), 0),
        # Fungsi segitiga untuk 'sedang' (20k → 40k)
        'sedang': max(min((x - 20000) / 10000, (40000 - x) / 10000), 0),
        # Fungsi segitiga naik untuk 'mahal' (30k → 50k+)
        'mahal': max(min((x - 30000) / 10000, 1), 0)
    }

# Fungsi untuk variabel output KELAYAKAN (0-100)
def kelayakan_membership(x):
    return {
        # Mirip dengan servis karena sama-sama skala 0-100
        'rendah': max(min((50 - x) / 30, 1), 0),
        'sedang': max(min((x - 40) / 20, (80 - x) / 20), 0),
        'tinggi': max(min((x - 60) / 30, 1), 0)
    }

# ================= SISTEM INFERENSI FUZZY =================
def fuzzy_rules(servis, harga):
    rules = []
    # Kombinasi semua kondisi input
    for s_key, s_val in servis.items():
        for h_key, h_val in harga.items():
            if s_val > 0 and h_val > 0:
                # Menggunakan operator MIN untuk implikasi
                strength = min(s_val, h_val)
                
                # Rule-base menentukan output kelayakan
                if s_key == 'baik' and h_key == 'murah':
                    rules.append(('tinggi', strength))  # Servis baik + harga murah → kelayakan tinggi
                elif s_key == 'baik' and h_key == 'sedang':
                    rules.append(('tinggi', strength))  # Servis baik + harga sedang → kelayakan tinggi
                elif s_key == 'sedang' and h_key == 'murah':
                    rules.append(('sedang', strength))  # Servis sedang + harga murah → kelayakan sedang
                elif s_key == 'buruk' or h_key == 'mahal':
                    rules.append(('rendah', strength))  # Servis buruk ATAU harga mahal → kelayakan rendah
                else:
                    rules.append(('sedang', strength))  # Default case
    return rules

# ================= METODE DEFUZZIFIKASI =================
def defuzzifikasi(rules):
    numerator = 0.0
    denominator = 0.0
    
    # Metode Centroid: hitung luas area di bawah kurva
    for x in range(0, 101):  # Evaluasi setiap titik dari 0-100
        membership = 0.0
        # Cari nilai keanggotaan agregasi (MAX-MIN)
        for term, strength in rules:
            mf_value = kelayakan_membership(x)[term]
            membership = max(membership, min(strength, mf_value))
        
        # Hitung momen dan luas
        numerator += x * membership  # Untuk momen
        denominator += membership    # Untuk luas
    
    # Centroid = total momen / total luas
    return numerator / denominator if denominator != 0 else 0

# ================= PROSES UTAMA =================
def proses_fuzzy_excel(input_file, output_file):
    # 1. Baca data input
    wb = load_workbook(input_file)
    ws = wb.active
    
    # 2. Proses setiap restoran
    restaurants = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        id_resto, servis, harga = row
        
        # Fuzzifikasi input
        m_servis = servis_membership(servis)
        m_harga = harga_membership(harga)
        
        # Evaluasi rule-base
        rules = fuzzy_rules(m_servis, m_harga)
        
        # Defuzzifikasi
        skor = defuzzifikasi(rules)
        
        restaurants.append({
            'ID': id_resto,
            'Servis': servis,
            'Harga': harga,
            'Skor': skor
        })
    
    # Mengurutkan dan ambil 5 terbaik
    top_5 = sorted(restaurants, key=lambda x: x['Skor'], reverse=True)[:5]
    
    # Simpan output
    wb_output = Workbook()
    ws_output = wb_output.active
    # Header kolom
    ws_output.append(["ID Pelanggan", "Kualitas Servis (1-100)", "Harga (Rp)", "Skor Kelayakan (0-100)"])
    
    for resto in top_5:
        ws_output.append([
            resto['ID'],
            resto['Servis'],
            resto['Harga'],
            round(float(resto['Skor']), 4)  # Pembulatan 4 digit
        ])
    
    wb_output.save(output_file)
    print(f"Hasil peringkat 5 restoran terbaik disimpan di: {output_file}")

# Eksekusi program
if __name__ == "__main__":
    proses_fuzzy_excel("restoran.xlsx", "peringkat.xlsx")